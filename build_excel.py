"""
Convert the GameChanger raw scrape JSON into a multi-sheet Excel workbook.

Scope: ONLY Bumblebeers offensive at-bats. Opposing-team at-bats are dropped
from the AtBats/Pitches/BaseRunning sheets. PlaysRaw still contains everything
for archival.

Sheets:
  Teams         - one row per Bumblebee* team season
  Schedule      - one row per game (date, opponent, score, etc.)
  Players       - one row per player (across all team-seasons)
  AtBats        - one row per BUMBLEBEERS at-bat
  Pitches       - one row per BUMBLEBEERS pitch event (during BMBL at-bat)
  BaseRunning   - one row per base_running event during BMBL at-bats
  PlaysRaw      - one row per raw event (BOTH teams; full archive)
  Errors        - the scrape errors (mostly upcoming/cancelled games)
"""

from __future__ import annotations
import json
import os
from collections import defaultdict
from datetime import datetime
from typing import Any, Iterator
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
RAW_JSON = os.path.join(HERE, "gamechanger_bumblebeers_raw.json")
OUT_XLSX = os.path.join(HERE, "bumblebeers_gamechanger.xlsx")


def load() -> dict:
    with open(RAW_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def safe_dt(s):
    if not s:
        return None
    if isinstance(s, dict):
        s = s.get("datetime")
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).isoformat()
    except Exception:
        return s


# ---------------------------------------------------------------------------
# Player lookup (keyed across all teams the user has)
# ---------------------------------------------------------------------------

def build_player_index(data):
    idx = {}
    for t in data["teams"]:
        tm = t["team_meta"]
        for p in (t.get("players") or []):
            if not isinstance(p, dict):
                continue
            pid = p.get("id")
            if not pid:
                continue
            idx[pid] = {
                "first_name": p.get("first_name"),
                "last_name": p.get("last_name"),
                "number": p.get("number"),
                "team_id": tm["id"],
                "season_year": tm["season_year"],
                "team_name": tm["name"],
                "bats": p.get("bats"),
                "status": p.get("status"),
            }
    return idx


def player_name(idx, pid):
    if not pid:
        return ""
    p = idx.get(pid)
    if not p:
        return pid[:8]
    return f"{p.get('first_name') or ''} {p.get('last_name') or ''}".strip() or pid[:8]


# ---------------------------------------------------------------------------
# Game state walker - tracks the BUMBLEBEERS offensive half-innings only
# ---------------------------------------------------------------------------

_OUT_PLAY_RESULTS = {
    "batter_out",
    "batter_out_advance_runners",
    "sacrifice_fly",
    "infield_fly",
    "strike_out",
    "strike_out_swinging",
    "strike_out_looking",
    "out_caught_stealing",
    "out_pickoff",
}

_OUT_BASERUN_TYPES = {
    "out_on_last_play",
    "caught_stealing",
    "picked_off",
    "out_at_base",
    "out",
}


class GameState:
    """Walks a game stream tracking the Bumblebeers offensive innings.

    Approach:
    - Bumblebeers offense starts in top of inning 1 if BMBL is away, bottom of 1 if home.
    - We only emit at-bats during the BMBL offensive half.
    - BMBL offensive inning counter advances when:
        * an end_half event fires DURING BMBL offense (i.e., end of BMBL's frame)
        * OR 3 outs accumulate during BMBL offense (lazy fallback)
    - The script treats "Bumblebeers half-inning N" as a clean counter regardless
      of what the opposing team does in between. So `bmbl_frame` = 1,2,3...
    """

    def __init__(self, owning_team_id, home_away):
        self.owning_team_id = owning_team_id
        self.home_away = home_away  # "home" / "away" for owning team
        # Bumblebeers bat top if AWAY, bottom if HOME.
        self.bmbl_bats_top = (home_away != "home")
        # Track whose half-inning we're in: True = BMBL offense
        self.bmbl_offense = self.bmbl_bats_top  # they bat first if away
        self.bmbl_frame = 1
        self.outs = 0  # only meaningful during BMBL offense
        # Lineups across BOTH teams (keyed by team_id)
        self.lineups = defaultdict(dict)
        self.lineup_index = defaultdict(int)

    def current_batter(self):
        if not self.bmbl_offense:
            return None
        lineup = self.lineups[self.owning_team_id]
        if not lineup:
            return None
        # Wrap the lineup index around (typical 9-12 player slots).
        size = max(lineup.keys()) + 1
        idx = self.lineup_index.get(self.owning_team_id, 0) % size
        return lineup.get(idx)

    def _flip_offense(self):
        """End of BMBL half-inning. Reset outs, advance counter at next start."""
        if self.bmbl_offense:
            self.bmbl_frame += 1
        self.bmbl_offense = not self.bmbl_offense
        self.outs = 0

    def maybe_auto_flip(self, upcoming_code):
        """Lazy 3-out flip: trigger if outs >= 3 and next event isn't end_half.

        Only counts BMBL outs (scorers may not record opposing-team at-bats),
        so the flip only fires reliably for BMBL halves. Half-inning numbering
        is therefore best-effort — see notes in the docstring.
        """
        if self.outs >= 3 and upcoming_code != "end_half":
            self._flip_offense()

    def apply_event(self, ed):
        code = ed.get("code")
        attrs = ed.get("attributes") or {}
        if code == "end_half":
            self._flip_offense()
        elif code == "fill_lineup_index":
            tid = attrs.get("teamId"); i = attrs.get("index"); pid = attrs.get("playerId")
            if tid and i is not None and pid:
                self.lineups[tid][int(i)] = pid
        elif code == "fill_lineup":
            tid = attrs.get("teamId"); pid = attrs.get("playerId")
            if tid and pid:
                used = set(self.lineups[tid].keys())
                next_i = 0
                while next_i in used:
                    next_i += 1
                self.lineups[tid][next_i] = pid
        elif code == "goto_lineup_index":
            tid = attrs.get("teamId"); i = attrs.get("index")
            if tid and i is not None:
                self.lineup_index[tid] = int(i)
        elif code == "clear_lineup_index":
            tid = attrs.get("teamId"); i = attrs.get("index")
            if tid and i is not None and int(i) in self.lineups[tid]:
                del self.lineups[tid][int(i)]
        elif code == "clear_entire_lineup":
            tid = attrs.get("teamId")
            if tid:
                self.lineups[tid].clear()
        elif code == "sub_players":
            tid = attrs.get("teamId")
            out_pid = attrs.get("outgoingPlayerId"); in_pid = attrs.get("incomingPlayerId")
            if tid and out_pid and in_pid:
                for idx, pid in list(self.lineups[tid].items()):
                    if pid == out_pid:
                        self.lineups[tid][idx] = in_pid

    def register_at_bat(self, play_result, base_running_inner):
        """After a transaction is a completed at-bat.

        We always count outs (so the 3-out fallback also flips Mets-frame halves)
        but only advance the BMBL lineup index during BMBL offense.
        """
        if self.bmbl_offense:
            self.lineup_index[self.owning_team_id] = self.lineup_index.get(self.owning_team_id, 0) + 1
        new_outs = 0
        if play_result in _OUT_PLAY_RESULTS:
            new_outs += 1
        for br in base_running_inner:
            pt = (br.get("attributes") or {}).get("playType")
            if pt in _OUT_BASERUN_TYPES:
                new_outs += 1
        self.outs += new_outs


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def teams_rows(data):
    rows = []
    for t in data["teams"]:
        tm = t["team_meta"]
        det = t.get("team_detail") or {}
        rec = (det.get("record") or {}) if isinstance(det, dict) else {}
        games_with_data = sum(1 for g in t["games"] if g.get("plays"))
        rows.append({
            "team_name": tm.get("name"),
            "season_year": tm.get("season_year"),
            "season_name": tm.get("season_name"),
            "age_group": tm.get("age_group"),
            "sport": tm.get("sport"),
            "city": tm.get("city"),
            "country": tm.get("country"),
            "archived": tm.get("archived"),
            "wins": rec.get("wins"),
            "losses": rec.get("losses"),
            "ties": rec.get("ties"),
            "games_scheduled": len(t["games"]),
            "games_with_data": games_with_data,
            "url_name": tm.get("url_encoded_name"),
            "public_id": tm.get("public_id"),
            "team_id": tm.get("id"),
            "url": f"https://web.gc.com/teams/{tm.get('public_id')}/{tm.get('url_encoded_name')}/schedule",
        })
    return rows


def schedule_rows(data):
    rows = []
    for t in data["teams"]:
        tm = t["team_meta"]
        gsums = {gs.get("event_id"): gs for gs in (t.get("game_summaries") or []) if isinstance(gs, dict)}
        for g in t["games"]:
            entry = g["schedule_entry"]
            ev = entry["event"]; pre = entry.get("pregame_data") or {}
            eid = ev.get("id")
            gs = gsums.get(eid, {})
            rows.append({
                "season_year": tm["season_year"],
                "team_name": tm["name"],
                "date_local": safe_dt(ev.get("start")),
                "opponent_name": pre.get("opponent_name"),
                "home_away": pre.get("home_away"),
                "bmbl_score": gs.get("owning_team_score"),
                "opponent_score": gs.get("opponent_team_score"),
                "game_status": gs.get("game_status") or ev.get("status"),
                "innings_played": (gs.get("sport_specific", {}).get("bats", {}).get("inning_details") or {}).get("inning"),
                "location": (ev.get("location") or {}).get("name") if isinstance(ev.get("location"), dict) else None,
                "title": ev.get("title"),
                "plays_count": len(g.get("plays") or []) if g.get("plays") else 0,
                "event_id": eid,
                "game_stream_id": g.get("game_stream_id"),
                "team_id": tm["id"],
                "url": f"https://web.gc.com/teams/{tm.get('public_id')}/{tm.get('url_encoded_name')}/schedule/{eid}/plays" if eid else None,
            })
    rows.sort(key=lambda r: (r["date_local"] or ""))
    return rows


def players_rows(data):
    rows = []
    for t in data["teams"]:
        tm = t["team_meta"]
        for p in (t.get("players") or []):
            if not isinstance(p, dict):
                continue
            rows.append({
                "season_year": tm["season_year"],
                "team_name": tm["name"],
                "first_name": p.get("first_name"),
                "last_name": p.get("last_name"),
                "number": p.get("number"),
                "bats": p.get("bats"),
                "status": p.get("status"),
                "player_id": p.get("id"),
            })
    return rows


def iter_plays(g):
    for p in g.get("plays") or []:
        try:
            ed = json.loads(p["event_data"])
        except Exception:
            continue
        yield p["sequence_number"], ed


def build_event_rows(data, players):
    atbats = []; pitches = []; baserun = []; raw_rows = []

    for t in data["teams"]:
        tm = t["team_meta"]
        sy = tm["season_year"]
        team_name = tm["name"]
        for g in t["games"]:
            entry = g["schedule_entry"]
            ev = entry["event"]; pre = entry.get("pregame_data") or {}
            eid = ev.get("id")
            date_str = safe_dt(ev.get("start"))
            opponent_name = pre.get("opponent_name")
            home_away = pre.get("home_away")

            gs = GameState(owning_team_id=tm["id"], home_away=home_away)

            def name(pid):
                return player_name(players, pid) if pid else ""

            for seq, ed in iter_plays(g):
                code = ed.get("code")
                gs.maybe_auto_flip(code)

                # Raw archive row (all events, both teams' frames)
                raw_rows.append({
                    "season_year": sy,
                    "date_local": date_str,
                    "opponent": opponent_name,
                    "event_id": eid,
                    "sequence_number": seq,
                    "code": code,
                    "in_bmbl_half": gs.bmbl_offense,
                    "bmbl_frame": gs.bmbl_frame,
                    "event_data": json.dumps(ed, separators=(",", ":")),
                })

                if code == "transaction":
                    # Snapshot before processing inner events
                    inning_before = gs.bmbl_frame
                    bmbl_half = gs.bmbl_offense
                    batter_id_before = gs.current_batter() if bmbl_half else None

                    inner_pitches = []
                    play_result = None
                    extended_play_result = None
                    play_type = None
                    defender_position = None
                    defender_x = None
                    defender_y = None
                    had_error = False
                    inner_baserun = []

                    for sub in ed.get("events", []) or []:
                        sc = sub.get("code"); sa = sub.get("attributes") or {}
                        if sc == "pitch":
                            inner_pitches.append(sa.get("result"))
                            if bmbl_half:
                                pitches.append({
                                    "season_year": sy,
                                    "date_local": date_str,
                                    "opponent": opponent_name,
                                    "event_id": eid,
                                    "transaction_seq": seq,
                                    "bmbl_frame": inning_before,
                                    "batter": name(batter_id_before),
                                    "batter_id": batter_id_before,
                                    "result": sa.get("result"),
                                    "advances_count": sa.get("advancesCount"),
                                    "advances_runners": sa.get("advancesRunners"),
                                })
                        elif sc == "ball_in_play":
                            play_result = sa.get("playResult")
                            extended_play_result = sa.get("extendedPlayResult")
                            play_type = sa.get("playType")
                            defenders = sa.get("defenders") or []
                            if defenders:
                                df = defenders[0]
                                defender_position = df.get("position")
                                if isinstance(df.get("location"), dict):
                                    defender_x = df["location"].get("x")
                                    defender_y = df["location"].get("y")
                                if df.get("error"):
                                    had_error = True
                        elif sc == "base_running":
                            inner_baserun.append(sub)
                            if bmbl_half:
                                baserun.append({
                                    "season_year": sy,
                                    "date_local": date_str,
                                    "opponent": opponent_name,
                                    "event_id": eid,
                                    "transaction_seq": seq,
                                    "bmbl_frame": inning_before,
                                    "runner": name(sa.get("runnerId")),
                                    "runner_id": sa.get("runnerId"),
                                    "play_type": sa.get("playType"),
                                    "base": sa.get("base"),
                                })
                        if sc in ("fill_lineup_index", "fill_lineup", "goto_lineup_index",
                                  "clear_lineup_index", "clear_entire_lineup", "sub_players"):
                            gs.apply_event(sub)

                    is_atbat = bool(inner_pitches and (play_result or extended_play_result))
                    if is_atbat and bmbl_half:
                        atbats.append({
                            "season_year": sy,
                            "date_local": date_str,
                            "opponent": opponent_name,
                            "event_id": eid,
                            "transaction_seq": seq,
                            "bmbl_frame": inning_before,
                            "batter": name(batter_id_before),
                            "batter_id": batter_id_before,
                            "n_pitches": len(inner_pitches),
                            "pitch_results": "|".join([r or "" for r in inner_pitches]),
                            "result": play_result,
                            "extended_result": extended_play_result,
                            "play_type": play_type,
                            "defender_position": defender_position,
                            "defender_x": defender_x,
                            "defender_y": defender_y,
                            "fielding_error": had_error,
                            "n_baserunning_events": len(inner_baserun),
                        })
                    if is_atbat:
                        # Always track outs/lineup advance — even in opposing frames
                        # so the BMBL state remains consistent.
                        gs.register_at_bat(play_result or extended_play_result, inner_baserun)
                else:
                    gs.apply_event(ed)

    return atbats, pitches, baserun, raw_rows


def errors_rows(data):
    return [{"event_id": e.get("eid"), "step": e.get("step"), "status": e.get("status")} for e in data.get("errors", [])]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    data = load()
    print(f"loaded scrape from {data.get('scraped_at')}")
    print(f"  teams: {len(data['teams'])}")

    players = build_player_index(data)
    print(f"  players indexed: {len(players)}")

    teams = teams_rows(data)
    schedule = schedule_rows(data)
    player_list = players_rows(data)
    atbats, pitches, baserun, raw_rows = build_event_rows(data, players)
    errors = errors_rows(data)

    print("  rows:")
    print(f"    Teams        {len(teams)}")
    print(f"    Schedule     {len(schedule)}")
    print(f"    Players      {len(player_list)}")
    print(f"    AtBats (BMBL)        {len(atbats)}")
    print(f"    Pitches (BMBL)       {len(pitches)}")
    print(f"    BaseRunning (BMBL)   {len(baserun)}")
    print(f"    PlaysRaw     {len(raw_rows)}")
    print(f"    Errors       {len(errors)}")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(teams).to_excel(writer, sheet_name="Teams", index=False)
        pd.DataFrame(schedule).to_excel(writer, sheet_name="Schedule", index=False)
        pd.DataFrame(player_list).to_excel(writer, sheet_name="Players", index=False)
        pd.DataFrame(atbats).to_excel(writer, sheet_name="AtBats", index=False)
        pd.DataFrame(pitches).to_excel(writer, sheet_name="Pitches", index=False)
        pd.DataFrame(baserun).to_excel(writer, sheet_name="BaseRunning", index=False)
        pd.DataFrame(raw_rows).to_excel(writer, sheet_name="PlaysRaw", index=False)
        pd.DataFrame(errors).to_excel(writer, sheet_name="Errors", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb = load_workbook(OUT_XLSX)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True)
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for c in col_cells[:200]:
                v = c.value
                if v is None:
                    continue
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)
    wb.save(OUT_XLSX)

    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
