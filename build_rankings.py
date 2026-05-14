"""
BMBL+ rankings — compute composite offensive scores at three granularities:

  - Per game (from play-by-play in gamechanger_bumblebeers_raw.json)
  - Per season (from gamechanger_season_stats.json)
  - Per career (two methods: PA-weighted average of season scores, AND
                recomputed on lifetime totals)

Outputs:
  - bumblebeers_rankings.xlsx  — multi-sheet workbook (Summary, Components,
                                 YearByYear, Career, Reconciliation, RawStats,
                                 PerGame, Weights)
  - bumblebeers_rankings.html  — interactive trending viewer (companion script
                                 build_rankings_html.py writes this)
"""

from __future__ import annotations
import json
import os
import math
from collections import defaultdict
from datetime import datetime
import pandas as pd
import numpy as np

HERE = os.path.dirname(os.path.abspath(__file__))
SEASON_JSON = os.path.join(HERE, "gamechanger_season_stats.json")
RAW_JSON = os.path.join(HERE, "gamechanger_bumblebeers_raw.json")
OUT_XLSX = os.path.join(HERE, "bumblebeers_rankings.xlsx")
PERGAME_JSON = os.path.join(HERE, "_pergame.json")  # intermediate for HTML viewer

# ---------------------------------------------------------------------------
# BMBL+ weights (edit these to retune)
# ---------------------------------------------------------------------------
WEIGHTS = {
    "wOBA":             0.40,   # overall production
    "ISO":              0.10,   # power
    "RISP_diff":        0.10,   # clutch - RISP performance vs own baseline
    "TwoOutRBI_rate":   0.08,   # clutch - 2-out RBI per PA
    "ProductiveOut":    0.07,   # clutch - sac flies + advance-the-runner outs
    "K_avoid":          0.05,   # discipline - 1 - K/PA
    "BB_rate":          0.05,   # discipline - BB/PA
    "QAB_pct":          0.05,   # discipline - quality at-bats
    "Hard_pct":         0.06,   # contact quality - hard hits
    "LD_pct":           0.04,   # contact quality - line drives
}
assert abs(sum(WEIGHTS.values()) - 1.0) < 1e-6, "Weights must sum to 1.0"

MIN_PA_QUALIFIED = 25
SHRINKAGE_K = 50            # PAs of "league mean" added for Bayesian shrinkage
SCORE_CENTER = 100
SCORE_STDDEV_POINTS = 25    # 1 stddev = 25 BMBL+ points

# Linear-weights for wOBA (close to MLB but tuned a bit for slo-pitch)
WOBA_WEIGHTS = {
    "BB": 0.69, "HBP": 0.72, "1B": 0.88, "2B": 1.25, "3B": 1.58, "HR": 2.10,
}

# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def load_data():
    with open(SEASON_JSON, "r", encoding="utf-8") as f:
        season = json.load(f)
    with open(RAW_JSON, "r", encoding="utf-8") as f:
        raw = json.load(f)
    return season, raw

# ---------------------------------------------------------------------------
# Player index across seasons
# ---------------------------------------------------------------------------

def build_player_lookup(season_data, raw_data):
    """Map player_id -> {first_name, last_name, number, display_name}.

    Uses both season-stats roster and raw scrape roster for coverage.
    """
    out = {}
    sources = []
    for s in season_data["teams"]:
        for p in (s.get("players") or []):
            if isinstance(p, dict): sources.append(p)
    for t in raw_data["teams"]:
        for p in (t.get("players") or []):
            if isinstance(p, dict): sources.append(p)
    for p in sources:
        pid = p.get("id")
        if not pid or pid in out:
            continue
        fn = (p.get("first_name") or "").strip()
        ln = (p.get("last_name") or "").strip()
        num = p.get("number")
        out[pid] = {
            "first_name": fn,
            "last_name": ln,
            "number": num,
            "display_name": (fn + (" " + ln if ln else "")).strip() or pid[:8],
        }
    return out

# ---------------------------------------------------------------------------
# Season-stats DataFrame
# ---------------------------------------------------------------------------

def build_season_df(season_data, players):
    """Flatten season-stats into a tidy DataFrame, one row per (player, season)."""
    rows = []
    for t in season_data["teams"]:
        tm = t["team_meta"]
        season_year = tm.get("season_year")
        sstats = t.get("season_stats") or {}
        data = (sstats.get("stats_data") or {}).get("players") or {}
        for pid, blob in data.items():
            off = (blob.get("stats") or {}).get("offense") or {}
            if not off:
                continue
            r = {"player_id": pid, "season_year": season_year, "team_name": tm.get("name")}
            r.update({k: off.get(k) for k in off.keys()})
            p = players.get(pid, {})
            r["display_name"] = p.get("display_name") or pid[:8]
            r["number"] = p.get("number")
            rows.append(r)
    df = pd.DataFrame(rows)
    df = df.sort_values(["season_year", "display_name"]).reset_index(drop=True)
    # Build a stable cross-season person_key from name (GameChanger creates
    # a fresh player_id per team-season, so we'd otherwise treat the same
    # human as a different player every year).
    # Manual merges confirmed by user:
    NAME_ALIASES = {
        "alex tosun": "alex",
        "brandon porco": "porco",
        "z.terence": "terence",
        # "Chris USPL" stays separate from "Chris" per user
    }
    def _pk(row):
        dn = (row.get("display_name") or "").strip().lower()
        return NAME_ALIASES.get(dn, dn)
    df["person_key"] = df.apply(_pk, axis=1)
    return df

# ---------------------------------------------------------------------------
# Component computations
# ---------------------------------------------------------------------------

def add_components_season(df: pd.DataFrame) -> pd.DataFrame:
    """Compute the raw per-(player, season) component values that feed BMBL+."""
    g = lambda col, default=0: df[col].fillna(default) if col in df.columns else default

    PA = g("PA"); AB = g("AB"); H = g("H")
    BB = g("BB"); HBP = g("HBP"); SO = g("SO")
    _1B = g("1B"); _2B = g("2B"); _3B = g("3B"); HR = g("HR")
    SF = g("SHF"); SH = g("SHB")
    AVG = g("AVG"); OBP = g("OBP"); SLG = g("SLG")
    ABRISP = g("ABRISP"); HRISP = g("HRISP"); BARISP = g("BA/RISP")
    TWO_OUT_RBI = g("2OUTRBI")
    QAB = g("QAB")
    HARD = g("HARD"); WEAK = g("WEAK"); LND = g("LND"); INP = g("INP")

    # wOBA denominator includes AB+BB+HBP+SF (standard formula)
    denom_wOBA = (AB + BB + HBP + SF).replace(0, np.nan)
    df["wOBA_raw"] = (
        WOBA_WEIGHTS["BB"] * BB +
        WOBA_WEIGHTS["HBP"] * HBP +
        WOBA_WEIGHTS["1B"] * _1B +
        WOBA_WEIGHTS["2B"] * _2B +
        WOBA_WEIGHTS["3B"] * _3B +
        WOBA_WEIGHTS["HR"] * HR
    ) / denom_wOBA

    df["ISO"] = (SLG - AVG)
    # RISP differential — only meaningful with some RISP at-bats
    risp_avg = (HRISP / ABRISP.replace(0, np.nan))
    df["RISP_diff"] = (risp_avg - AVG).where(ABRISP >= 5)  # noise floor

    df["TwoOutRBI_rate"] = (TWO_OUT_RBI / PA.replace(0, np.nan))
    df["ProductiveOut"]  = ((SF + SH) / PA.replace(0, np.nan))
    df["K_avoid"]        = 1 - (SO / PA.replace(0, np.nan))
    df["BB_rate"]        = (BB / PA.replace(0, np.nan))
    df["QAB_pct"]        = (QAB / PA.replace(0, np.nan))

    contact = (HARD + WEAK).replace(0, np.nan)
    df["Hard_pct"] = (HARD / contact)
    df["LD_pct"]   = (LND / INP.replace(0, np.nan))
    return df

def bayesian_shrink_wOBA(df: pd.DataFrame, k: int = SHRINKAGE_K) -> pd.DataFrame:
    """Regress each player's wOBA toward season mean using k pseudo-PA."""
    df = df.copy()
    df["wOBA_shrunk"] = np.nan
    for season, grp in df.groupby("season_year"):
        # Weighted team-mean wOBA by PA
        mask = grp["PA"].fillna(0) > 0
        if not mask.any():
            continue
        season_wOBA = (grp["wOBA_raw"] * grp["PA"]).sum() / grp["PA"].sum()
        eff = (grp["PA"] * grp["wOBA_raw"] + k * season_wOBA) / (grp["PA"] + k)
        df.loc[grp.index, "wOBA_shrunk"] = eff
    return df

def add_zscores_and_score(df: pd.DataFrame) -> pd.DataFrame:
    """For each season, z-score each component within the season, then sum-weight."""
    components = list(WEIGHTS.keys())
    # Use shrunk wOBA as the production input
    df["wOBA"] = df["wOBA_shrunk"]
    out = df.copy()
    for c in components:
        out[c + "_z"] = np.nan
    out["BMBL_plus"] = np.nan
    out["qualified"] = out["PA"].fillna(0) >= MIN_PA_QUALIFIED

    for season, grp in out.groupby("season_year"):
        qual = grp[grp["qualified"]]
        if len(qual) < 2:
            continue
        for c in components:
            mu = qual[c].mean()
            sd = qual[c].std(ddof=0)
            if not sd or math.isnan(sd) or sd == 0:
                out.loc[grp.index, c + "_z"] = 0
            else:
                out.loc[grp.index, c + "_z"] = (grp[c] - mu) / sd
        weighted = sum(out.loc[grp.index, c + "_z"].fillna(0) * WEIGHTS[c] for c in components)
        out.loc[grp.index, "BMBL_plus"] = SCORE_CENTER + SCORE_STDDEV_POINTS * weighted
    out["season_rank"] = (
        out[out["qualified"]]
          .groupby("season_year")["BMBL_plus"]
          .rank(method="min", ascending=False)
    )
    return out

# ---------------------------------------------------------------------------
# Career roll-up (two methods)
# ---------------------------------------------------------------------------

def career_method_a(season_df: pd.DataFrame) -> pd.DataFrame:
    """PA-weighted average of qualified season BMBL+. Keyed by person_key (name)."""
    q = season_df[season_df["qualified"] & season_df["BMBL_plus"].notna()].copy()
    if q.empty:
        return pd.DataFrame()
    rows = []
    for pk, x in q.groupby("person_key"):
        rows.append({
            "person_key": pk,
            "display_name": x["display_name"].iloc[0],
            "seasons_qualified": len(x),
            "career_PA": int(x["PA"].sum()),
            "career_BMBLplus_weighted": float((x["BMBL_plus"] * x["PA"]).sum() / x["PA"].sum()),
            "seasons_played": ", ".join(str(int(y)) for y in sorted(x["season_year"].unique())),
            "peak_season_year": int(x.loc[x["BMBL_plus"].idxmax(), "season_year"]),
            "peak_BMBLplus": round(float(x["BMBL_plus"].max()), 1),
        })
    agg = pd.DataFrame(rows).sort_values("career_BMBLplus_weighted", ascending=False)
    agg["career_BMBLplus_weighted"] = agg["career_BMBLplus_weighted"].round(1)
    return agg.reset_index(drop=True)

def career_method_b(season_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate lifetime totals, recompute components and BMBL+ across the whole league."""
    sum_cols = ["PA","AB","H","1B","2B","3B","HR","BB","HBP","SO","SHF","SHB",
                "ABRISP","HRISP","2OUTRBI","QAB","HARD","WEAK","LND","INP","TB"]
    df = season_df.copy()
    df = df[df["PA"].fillna(0) > 0]
    cols_present = [c for c in sum_cols if c in df.columns]
    agg = df.groupby(["player_id","display_name"])[cols_present].sum().reset_index()
    # Derive AVG/SLG/OBP/etc. from totals
    agg["AVG"] = agg["H"] / agg["AB"].replace(0, np.nan)
    agg["OBP"] = (agg["H"] + agg["BB"] + agg["HBP"]) / (agg["AB"] + agg["BB"] + agg["HBP"] + agg["SHF"]).replace(0, np.nan)
    agg["SLG"] = agg["TB"] / agg["AB"].replace(0, np.nan)
    agg["OPS"] = agg["OBP"] + agg["SLG"]
    agg["season_year"] = "career"
    add_components_season(agg)
    # No shrinkage for career (large samples)
    agg["wOBA_shrunk"] = agg["wOBA_raw"]
    agg["wOBA"] = agg["wOBA_shrunk"]
    # Z-score within career population
    components = list(WEIGHTS.keys())
    qual_mask = agg["PA"] >= MIN_PA_QUALIFIED * 2  # require ≥50 career PA to "qualify" career-wise
    qual = agg[qual_mask]
    for c in components:
        if c in qual.columns:
            mu = qual[c].mean()
            sd = qual[c].std(ddof=0)
            agg[c + "_z"] = 0 if (not sd or math.isnan(sd)) else (agg[c] - mu) / sd
    agg["career_BMBLplus_totals"] = SCORE_CENTER + SCORE_STDDEV_POINTS * sum(
        agg[c + "_z"].fillna(0) * WEIGHTS[c] for c in components
    )
    agg["qualified_career"] = qual_mask
    return agg[["player_id", "display_name", "career_PA" if "career_PA" in agg.columns else "PA",
                "AVG", "OBP", "SLG", "OPS", "wOBA", "ISO",
                "career_BMBLplus_totals", "qualified_career"]].rename(columns={"PA": "career_PA"})

# ---------------------------------------------------------------------------
# Per-game wOBA from raw at-bats (for trending viewer)
# ---------------------------------------------------------------------------

_OUT_RESULTS = {"batter_out", "batter_out_advance_runners", "sacrifice_fly",
                "infield_fly", "fielders_choice", "other_out",
                "dropped_third_strike_batter_out", "strike_out"}

def _outcome_counts(result):
    """Return dict of counts (1B, 2B, 3B, HR, BB, HBP, SO, etc.) for one at-bat result."""
    d = {"AB":0,"H":0,"1B":0,"2B":0,"3B":0,"HR":0,"BB":0,"HBP":0,"SO":0,"SHF":0,"SHB":0,"PA":1}
    if result == "single":           d.update({"AB":1, "H":1, "1B":1})
    elif result == "double":         d.update({"AB":1, "H":1, "2B":1})
    elif result == "triple":         d.update({"AB":1, "H":1, "3B":1})
    elif result == "home_run":       d.update({"AB":1, "H":1, "HR":1})
    elif result == "sacrifice_fly":  d.update({"SHF":1})
    elif result == "walk":           d.update({"BB":1})
    elif result == "hit_by_pitch":   d.update({"HBP":1})
    elif result == "strike_out":     d.update({"AB":1, "SO":1})
    elif result and result in _OUT_RESULTS: d.update({"AB":1})
    else:                            d.update({"AB":1})
    return d

def build_per_game_df(raw_data, players, season_df):
    """One row per (player, game) with totals + per-game wOBA."""
    rows = []
    # Index season wOBA for shrinkage
    season_wOBA = season_df.set_index(["player_id","season_year"])["wOBA"].to_dict()
    season_team_wOBA = season_df.groupby("season_year").apply(
        lambda x: (x["wOBA"]*x["PA"]).sum()/x["PA"].sum() if x["PA"].sum() else np.nan
    ).to_dict()

    for t in raw_data["teams"]:
        tm = t["team_meta"]
        sy = tm["season_year"]
        for g in t["games"]:
            ev = g["schedule_entry"]["event"]
            date = (ev.get("start") or {}).get("datetime")
            opp = (g["schedule_entry"].get("pregame_data") or {}).get("opponent_name")
            eid = ev.get("id")
            # Aggregate per-player counts for this game
            per_player = defaultdict(lambda: defaultdict(int))
            for p in (g.get("plays") or []):
                try:
                    ed = json.loads(p["event_data"])
                except: continue
                if ed.get("code") != "transaction":
                    continue
                pid = None
                result = None
                for sub in ed.get("events", []):
                    if sub.get("code") == "ball_in_play":
                        result = (sub.get("attributes") or {}).get("playResult")
                # we don't have per-AB batter_id reliably, so we'll skip those that the
                # script's earlier walker would have but use a separate aggregation:
                # (per-game stats by batter are loaded from the Excel AtBats sheet instead)
                pass
            # NOTE: per-game per-player batter resolution lives in the AtBats sheet we
            # already wrote out — read that file and aggregate by (date, batter, opponent).
    return rows  # placeholder; real aggregation below

def per_game_from_atbats(season_df):
    """Easier path: read AtBats sheet from the prior workbook and aggregate."""
    src = os.path.join(HERE, "bumblebeers_gamechanger.xlsx")
    if not os.path.exists(src):
        print("WARN: bumblebeers_gamechanger.xlsx not found; per-game disabled")
        return pd.DataFrame()
    ab = pd.read_excel(src, sheet_name="AtBats")
    ab["season_year"] = ab["season_year"].astype(int)
    # Map result -> counts
    counts = ab["result"].apply(_outcome_counts).apply(pd.Series)
    df = pd.concat([ab[["season_year","date_local","opponent","event_id","batter","batter_id"]], counts], axis=1)
    grouped = df.groupby(["season_year","date_local","opponent","event_id","batter","batter_id"]).sum(numeric_only=True).reset_index()
    # Compute per-game raw wOBA
    denom = grouped["AB"] + grouped["BB"] + grouped["HBP"] + grouped["SHF"]
    grouped["wOBA_raw_game"] = (
        WOBA_WEIGHTS["BB"]*grouped["BB"] + WOBA_WEIGHTS["HBP"]*grouped["HBP"] +
        WOBA_WEIGHTS["1B"]*grouped["1B"] + WOBA_WEIGHTS["2B"]*grouped["2B"] +
        WOBA_WEIGHTS["3B"]*grouped["3B"] + WOBA_WEIGHTS["HR"]*grouped["HR"]
    ) / denom.replace(0, np.nan)
    # Shrink each game toward player's season wOBA (or team wOBA if missing)
    season_lookup = season_df.set_index(["player_id","season_year"])["wOBA"].to_dict()
    team_season = season_df.groupby("season_year").apply(
        lambda x: (x["wOBA"]*x["PA"]).sum()/x["PA"].sum() if x["PA"].sum() else np.nan
    ).to_dict()
    def shrink(row):
        if pd.isna(row["wOBA_raw_game"]):
            return np.nan
        season = row["season_year"]
        pid = row["batter_id"]
        ref = season_lookup.get((pid, season)) or team_season.get(season)
        if ref is None or pd.isna(ref):
            return row["wOBA_raw_game"]
        return (row["PA"] * row["wOBA_raw_game"] + SHRINKAGE_K * ref) / (row["PA"] + SHRINKAGE_K)
    grouped["wOBA_shrunk_game"] = grouped.apply(shrink, axis=1)
    grouped["date_only"] = pd.to_datetime(grouped["date_local"]).dt.tz_localize(None).dt.date.astype(str)
    return grouped.sort_values(["batter","date_local"]).reset_index(drop=True)

# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

def reconcile(season_df, atbats_path):
    """Compare season-stats H/HR/etc. against play-by-play counts. Flag big gaps."""
    if not os.path.exists(atbats_path):
        return pd.DataFrame()
    ab = pd.read_excel(atbats_path, sheet_name="AtBats")
    counts = ab["result"].apply(_outcome_counts).apply(pd.Series)
    df = pd.concat([ab[["season_year","batter_id"]], counts], axis=1)
    pbp = df.groupby(["season_year","batter_id"]).sum(numeric_only=True).reset_index()
    pbp = pbp.rename(columns={"H":"pbp_H","HR":"pbp_HR","AB":"pbp_AB","PA":"pbp_PA","2B":"pbp_2B","3B":"pbp_3B"})
    merged = season_df.merge(pbp, left_on=["player_id","season_year"], right_on=["batter_id","season_year"], how="left")
    merged["H_diff_pct"]  = ((merged["pbp_H"]  - merged["H"])  / merged["H"].replace(0, np.nan)).abs() * 100
    merged["HR_diff_pct"] = ((merged["pbp_HR"] - merged["HR"]) / merged["HR"].replace(0, np.nan)).abs() * 100
    merged["AB_diff_pct"] = ((merged["pbp_AB"] - merged["AB"]) / merged["AB"].replace(0, np.nan)).abs() * 100
    cols = ["season_year","display_name","PA","AB","H","HR","pbp_PA","pbp_AB","pbp_H","pbp_HR",
            "H_diff_pct","HR_diff_pct","AB_diff_pct"]
    keep = [c for c in cols if c in merged.columns]
    return merged[keep].copy()

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    season_data, raw_data = load_data()
    players = build_player_lookup(season_data, raw_data)
    print(f"players: {len(players)}")

    season_df = build_season_df(season_data, players)
    print(f"player-seasons: {len(season_df)}")

    season_df = add_components_season(season_df)
    season_df = bayesian_shrink_wOBA(season_df)
    season_df = add_zscores_and_score(season_df)

    career_a = career_method_a(season_df)
    career_b = career_method_b(season_df)

    pergame_df = per_game_from_atbats(season_df)
    recon_df = reconcile(season_df, os.path.join(HERE, "bumblebeers_gamechanger.xlsx"))

    # Build sheets
    components = list(WEIGHTS.keys())
    summary_cols = ["season_year","display_name","number","PA","AB","H","HR","RBI","AVG","OBP","SLG","OPS",
                    "BA/RISP","2OUTRBI","QAB%","BMBL_plus","season_rank","qualified"]
    summary_cols = [c for c in summary_cols if c in season_df.columns]
    summary = season_df[summary_cols].copy()
    summary["BMBL_plus"] = summary["BMBL_plus"].round(1)

    comp_cols = ["season_year","display_name","PA"] + components + [c+"_z" for c in components] + ["BMBL_plus","season_rank"]
    comp_cols = [c for c in comp_cols if c in season_df.columns]
    components_sheet = season_df[comp_cols].copy()

    yearbyyear_frames = []
    for y, grp in season_df.groupby("season_year"):
        top10 = grp[grp["qualified"]].sort_values("BMBL_plus", ascending=False).head(10).copy()
        top10["BMBL_plus"] = top10["BMBL_plus"].round(1)
        top10 = top10[["season_year","display_name","PA","AVG","OBP","SLG","OPS","BMBL_plus","season_rank"]]
        yearbyyear_frames.append(top10)
    yearbyyear = pd.concat(yearbyyear_frames, ignore_index=True) if yearbyyear_frames else pd.DataFrame()

    weights_sheet = pd.DataFrame([
        {"component": k, "weight": v, "explanation": ""} for k,v in WEIGHTS.items()
    ])

    raw_stats = season_df.copy()

    # Write
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        summary.to_excel(w, "Summary", index=False)
        components_sheet.to_excel(w, "Components", index=False)
        yearbyyear.to_excel(w, "YearByYear", index=False)
        career_a.to_excel(w, "Career_Weighted", index=False)
        career_b.to_excel(w, "Career_Totals", index=False)
        recon_df.to_excel(w, "Reconciliation", index=False)
        pergame_df.to_excel(w, "PerGame", index=False)
        weights_sheet.to_excel(w, "Weights", index=False)
        raw_stats.to_excel(w, "RawStats", index=False)

    # Cosmetic formatting
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    wb = load_workbook(OUT_XLSX)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True)
        for col_cells in ws.columns:
            mx = 0; letter = col_cells[0].column_letter
            for c in col_cells[:200]:
                if c.value is None: continue
                mx = max(mx, len(str(c.value)))
            ws.column_dimensions[letter].width = min(max(10, mx+2), 40)
    wb.save(OUT_XLSX)

    # Dump per-game JSON for the HTML viewer
    if not pergame_df.empty:
        records = pergame_df.to_dict(orient="records")
        with open(PERGAME_JSON, "w", encoding="utf-8") as f:
            json.dump({
                "weights": WEIGHTS,
                "season_summaries": season_df[["player_id","display_name","season_year","PA","wOBA","BMBL_plus","qualified"]].to_dict(orient="records"),
                "career_weighted": career_a.to_dict(orient="records"),
                "career_totals": career_b.to_dict(orient="records"),
                "per_game": records,
            }, f, default=str)

    print(f"wrote {OUT_XLSX}")
    print(f"intermediate: {PERGAME_JSON}")

if __name__ == "__main__":
    main()
       if c.value is None: continue
                mx = max(mx, len(str(c.value)))
            ws.column_dimensions[letter].width = min(max(10, mx+2), 40)
    wb.save(OUT_XLSX)

    # Dump per-game JSON for the HTML viewer
    if not pergame_df.empty:
        records = pergame_df.to_dict(orient="records")
        with open(PERGAME_JSON, "w", encoding="utf-8") as f:
            json.dump({
                "weights": WEIGHTS,
                "season_summaries": season_df[["player_id","display_name","season_year","PA","wOBA","BMBL_plus","qualified"]].to_dict(orient="records"),
                "career_weighted": career_a.to_dict(orient="records"),
                "career_totals": career_b.to_dict(orient="records"),
                "per_game": records,
            }, f, default=str)

    print(f"wrote {OUT_XLSX}")
    print(f"intermediate: {PERGAME_JSON}")

if __name__ == "__main__":
    main()
